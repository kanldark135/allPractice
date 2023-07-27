#%% 1. 기본 순서
# 1) requests / BeautifulSoup from bs4
# 2) response = requests.get(url)
# 3) soup = BeautifulSoup(response.content, 'html.parser')
# 4) 이하 검색 등등

from bs4 import BeautifulSoup
import requests

url = 'https://davelee-fun.github.io/'
res = requests.get(url)
soup = BeautifulSoup(res.content, 'html.parser')

#%% 2. find 와 attrs 을 활용한 검색. find -> html 반환 / findall -> 리스트 번환

menu_list = soup.find('ul', attrs = {'class' : 'navbar-nav ml-auto'})
txt = menu_list.get_text()
link_list = [i.attrs['href']) for i in menu_list.select('a.nav-link')]

#%% 3. select 를 활용한 검색. select_one -> html 반환 / select -> 리스트 반환

subtitle = soup.select('div.mainheading > p.lead')[0].get_text().strip()

#%% 4. loop 로 뽑기 + elem.has_attr('href') 를 활용한 특정 속성 보유여부 체크

cat_elems = soup.select('section.featured-posts a.text-dark')
prod_elems = soup.select('section.featured-posts h4.card-text')

featured_prods = list()

for i, j in zip(cat_elems, prod_elems):
        if i.has_attr('href'):
                res = [i.get_text(), url + i.get('href'), j.get_text()]
                featured_prods.append(res)


#%% 5. if res.status_code != 200: 으로 코딩하는 관행 체크 

import requests
from bs4 import BeautifulSoup

res = requests.get("https://davelee-fun.github.io/index.html/xxx")

if res.status_code != 200:
       print ("error")

else:
      soup = BeautifulSoup(res.content, 'html.parser')

#%% 6. 여러 페이지 크롤링 동시에 -> requests.get(url1) -> / requests.get(url2)...


# 가령 동일 홈페이지에서도 페이지1 / 페이지2 식으로 있는 경우
# 아예 다른 여러 홈페이지도 이대로 확장 활용 가능

import requests
from bs4 import BeautifulSoup

for page_num in range(10):
    if page_num == 0:
        res = requests.get("https://davelee-fun.github.io/")
    else:
        res = requests.get("https://davelee-fun.github.io/page" + str(page_num + 1))
    
    soup = BeautifulSoup(res.content, 'html.parser')
    data = soup.select('h4.card-text')
    for item in data:
        print(item.get_text().strip())

#%% 7. to xls

import requests
from bs4 import BeautifulSoup
import openpyxl
import os
from datetime import datetime

os.chdir("C:/Users/문희관/Desktop")

# "엑셀로" 옮기는 함수

def data_input(filename, scraped_list, sheetname = None):

    xls_file = openpyxl.Workbook()
    xls_sheet = xls_file.active
    xls_sheet.column_dimensions['A'].width = 100

    if sheetname == None:
        xls_sheet.title = datetime.strftime(datetime.today(), '%Y-%m-%d')
    else:
         xls_sheet.title = str(sheetname)

    for i in scraped_list:
        xls_sheet.append(i)

    xls_file.save("./{0}.xlsx".format(filename))
    xls_file.close()

product_list = []

#####.....

for page_num in range(10):
    if page_num == 0:
        res = requests.get("https://davelee-fun.github.io/")
    else:
        res = requests.get("https://davelee-fun.github.io/page" + str(page_num + 1))
    
    soup = BeautifulSoup(res.content, 'html.parser')
    data = soup.select('div.card')

    for item in data:
        prod_names = item.select_one('h4.card-text')
        prod_date = item.select_one('div.wrapfooter span.post-date')
        prod_info = [prod_names.get_text().strip(), prod_date.get_text()]
        product_list.append(prod_info)
    
data_input('product', product_list)


#%% 8. from xlsx

import openpyxl

# 파일 열기
xls_file = openpyxl.load_workbook("C:/Users/문희관/Desktop/product.xlsx")

# 시트 열기
xls_file.sheetnames
xls_sheet = xls_file['2023-06-23']

# 시트 내 특정 셀 참조
iterrow = xls_sheet.rows # iterator 를 생성 후 LOOP 돌려서 추출

item_name = []
item_date = []
for i in iterrow:
    item_name.append(i[0].value)
    item_date.append(i[1].value)

xls_file.close()

# %% regex 로 찾기







# %% regex practice

import requests
import re
from bs4 import BeautifulSoup

res = requests.get("https://finance.naver.com/")
soup = BeautifulSoup(res.content, 'html.parser')

forex = soup.select_one('div.group2 table.tbl_home')

titles = [i.get_text() for i in forex.select("tbody th[scope]")]

item = forex.select('tbody td')
reg_pattern = re.compile('em class=\"b+')

chg = []
quote = []

for i in item:
    if re.search(reg_pattern, str(i)):
        chg.append(i.get_text())
    else:
        quote.append(i.get_text())

import pandas as pd

pd.DataFrame([chg, quote], columns = titles)

#%% 실습 2

import requests
from bs4 import BeautifulSoup

for sosok in range(1):
    if sosok == 0:
        res = requests.get("https://finance.naver.com/sise/sise_rise.naver?sosok=0")
    else:
        res = requests.get("https://finance.naver.com/sise/sise_rise.naver?sosok={0}".format(sosok))

    soup = BeautifulSoup(res.content, 'html.parser')

    tbl = soup.find('table', attrs = {'class' : 'type_2'})

    titles = tbl.select('tr')[0]
    company = []
    company_code = []
    current_price = []
    chg = []
    chg_percent = []

    for i in range(2, 100):
        row = tbl.select('tr')[i]
        if row.select_one('td.no'):
            company.append(row.select_one('td a.tltle').get_text())
            code = re.search("=[A-Za-z0-9]{6}", tbl.select('tr')[i].select_one('td a.tltle')['href']).group()[1:7]
            company_code.append(code)
            current_price.append(row.select('td')[2].get_text())
            chg.append(row.select('td')[3].get_text().strip())
            chg_percent.append(row.select('td')[4].get_text().strip())
        else:
            continue
        